import pandas as pd
from openpyxl import load_workbook
from payment_behavior_analysis import (find_the_last_day_current_month, calculate_outstanding_at_month_end)
from performance_review import calculate_overall_performance, calculate_performance_per_collector


def build_reports(customers, customer_to_collector, expected_dates, open_invoices,
                  closed_invoices, output_file="overdue_report.xlsx"):
    # Common filter
    last_day = find_the_last_day_current_month()
    filtered_invoices = [inv for inv in open_invoices if inv.due_date <= last_day]

    # --- First report: by overdue bucket ---
    df_bucket = pd.DataFrame([{
        "customer_name": inv.customer_name,
        "customer_id": inv.customer_id,
        "bucket": inv.bucket,
        "amount_usd": inv.amount_in_usd,
        "collector_name": customer_to_collector.get(inv.customer_id, ("Unknown", ""))[0],
        "collector_manager": customer_to_collector.get(inv.customer_id, ("Unknown", ""))[1]
    } for inv in filtered_invoices])

    report_bucket = df_bucket.pivot_table(
        index=["customer_name", "customer_id", "collector_name", "collector_manager"],
        columns="bucket",
        values="amount_usd",
        aggfunc="sum",
        fill_value=0
    )
    report_bucket["TotalUSD"] = report_bucket.sum(axis=1)
    report_bucket = report_bucket.reset_index()

    # --- Second report: by status ---
    df_status = pd.DataFrame([{
        "customer_name": inv.customer_name,
        "customer_id": inv.customer_id,
        "status": inv.status,
        "amount_usd": inv.amount_in_usd,
        "collector_name": customer_to_collector.get(inv.customer_id, ("Unknown", ""))[0],
    } for inv in filtered_invoices])

    report_status = df_status.pivot_table(
        index=["customer_name", "customer_id", "collector_name"],
        columns="status",
        values="amount_usd",
        aggfunc="sum",
        fill_value=0
    )
    report_status["TotalUSD"] = report_status.sum(axis=1)
    report_status = report_status.reset_index()

    # --- Third report: expected overdue on D ---
    outstanding_by_customer = calculate_outstanding_at_month_end(expected_dates, customers)

    df_outstanding = pd.DataFrame([{
        "customer_id": cust_id,
        "customer_name": next(
            (cust.customer_name for cust in customers if cust.payer_number == cust_id),
            "Unknown"
        ),
        "collector_name": customer_to_collector.get(cust_id, ("Unknown", ""))[0],
        "collector_manager": customer_to_collector.get(cust_id, ("Unknown", ""))[1],
        "ExpectedOverdueUSD": amount
    } for cust_id, amount in outstanding_by_customer.items()
        if amount > 0  # only include customers with outstanding balance
    ])

    report_expected = df_outstanding.pivot_table(
        index=["customer_name", "customer_id", "collector_name", "collector_manager"],
        values="ExpectedOverdueUSD",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # --- 4th report overall performance ---

    performance_result, clarification = calculate_overall_performance(open_invoices, closed_invoices)
    df_performance = pd.DataFrame([performance_result])

    # --- Fifth report: performance per collector ---
    performance_per_collector = calculate_performance_per_collector(
        open_invoices, closed_invoices, customer_to_collector
    )

    # Convert dict-of-dicts into DataFrame with collector name as a column
    df_perf_collector = pd.DataFrame([
        {"Collector Name": collector_id, **stats}
        for collector_id, stats in performance_per_collector.items()
    ])

    df_perf_collector = df_perf_collector.reset_index()

    # Reorder columns so Collector Name comes first
    df_perf_collector = df_perf_collector[
        ["Collector Name"] + [c for c in df_perf_collector.columns if c != "Collector Name"]]

    # --- Write all in the same Excel file ---
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        report_bucket.to_excel(writer, sheet_name="AgingReport", index=False)
        report_status.to_excel(writer, sheet_name="MonthEndPrepStatus", index=False)
        report_expected.to_excel(writer, sheet_name="ExpectedOverdueOnMonthEnd", index=False)
        df_performance.to_excel(writer, sheet_name="PerformanceSummary", index=False)
        df_perf_collector.to_excel(writer, sheet_name="PerformancePerCollector", index=False)

    # --- Append custom text to the third sheet ---
    wb = load_workbook(output_file)
    ws = wb["ExpectedOverdueOnMonthEnd"]
    last_row = ws.max_row + 2  # Leave a blank row
    ws.cell(row=last_row,
            column=1).value = "PLEASE FOCUS ON THESE CUSTOMERS. THERE IS A LATE PAYMENT EXPECTED THIS MONTH!"
    # --- Append clarifications to the fourth sheet ---
    ws_perf = wb["PerformanceSummary"]
    last_row_perf = ws_perf.max_row + 2

    row_number = last_row_perf  # start after the table

    # Write the header
    ws_perf.cell(row=row_number, column=1).value = "Clarifications:"

    # write each line one by one
    for line in clarification:
        row_number += 1  # move down one row
        ws_perf.cell(row=row_number, column=1).value = line

    # Save the file
    wb.save(output_file)

    print(f"Reports saved to {output_file} with 5 sheets.")
